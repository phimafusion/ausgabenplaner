import datetime
import io
import sqlite3
from typing import List, Dict, Any, Optional
from app.domain import calculate_plan_totals, format_currency_de
from app.backups import get_local_now



def get_version_details(conn: sqlite3.Connection, version_id: int) -> Optional[Dict[str, Any]]:
    ver_row = conn.execute("SELECT * FROM versions WHERE id = ?", (version_id,)).fetchone()
    if not ver_row:
        return None

    ver = dict(ver_row)

    # Fetch positions
    pos_rows = conn.execute(
        "SELECT * FROM positions WHERE version_id = ? ORDER BY sort_order ASC, id ASC", (version_id,)
    ).fetchall()
    positions = []
    for r in pos_rows:
        p = dict(r)
        p["amount_formatted"] = format_currency_de(p["amount"])
        positions.append(p)

    # Fetch contributions
    contrib_rows = conn.execute(
        "SELECT * FROM contributions WHERE version_id = ? ORDER BY sort_order ASC, id ASC", (version_id,)
    ).fetchall()
    contributions = []
    for r in contrib_rows:
        c = dict(r)
        c["amount_formatted"] = format_currency_de(c["amount"])
        contributions.append(c)

    totals = calculate_plan_totals(positions, contributions)
    ver["positions"] = positions
    ver["contributions"] = contributions
    ver["totals"] = totals
    return ver


def get_plan_details(conn: sqlite3.Connection, plan_id: int) -> Optional[Dict[str, Any]]:
    plan_row = conn.execute("SELECT * FROM plans WHERE id = ?", (plan_id,)).fetchone()
    if not plan_row:
        return None

    plan = dict(plan_row)
    plan["is_archived"] = bool(plan.get("is_archived", 0))

    ver_rows = conn.execute(
        "SELECT id, title, effective_date, is_active, created_at, created_by, updated_at, updated_by FROM versions WHERE plan_id = ? ORDER BY id DESC",
        (plan_id,),
    ).fetchall()
    versions = [dict(v) for v in ver_rows]
    plan["versions"] = versions

    # Active version is either marked active or newest
    active_ver_row = conn.execute("SELECT id FROM versions WHERE plan_id = ? AND is_active = 1 LIMIT 1", (plan_id,)).fetchone()
    if active_ver_row:
        plan["active_version"] = get_version_details(conn, active_ver_row["id"])
    elif versions:
        plan["active_version"] = get_version_details(conn, versions[0]["id"])
    else:
        plan["active_version"] = None

    return plan


def get_active_plan(conn: sqlite3.Connection, user_id: Optional[int] = None, user_role: str = "admin") -> Optional[Dict[str, Any]]:
    if user_role == "admin" or user_id is None:
        # First non-archived plan, fallback to any plan
        plan_row = conn.execute("SELECT id FROM plans WHERE is_archived = 0 ORDER BY id ASC LIMIT 1").fetchone()
        if not plan_row:
            plan_row = conn.execute("SELECT id FROM plans ORDER BY id ASC LIMIT 1").fetchone()
    else:
        # Check assigned plans
        assigned = conn.execute(
            """
            SELECT p.id FROM plans p
            JOIN user_plans up ON p.id = up.plan_id
            WHERE up.user_id = ? AND p.is_archived = 0
            ORDER BY p.id ASC LIMIT 1
            """,
            (user_id,),
        ).fetchone()
        if assigned:
            plan_row = assigned
        else:
            # Fallback if no specific assignment: first non-archived plan
            has_any_assignments = conn.execute("SELECT 1 FROM user_plans WHERE user_id = ?", (user_id,)).fetchone()
            if has_any_assignments:
                # User has assignments but all might be archived, try any assigned
                plan_row = conn.execute(
                    "SELECT p.id FROM plans p JOIN user_plans up ON p.id = up.plan_id WHERE up.user_id = ? ORDER BY p.id ASC LIMIT 1",
                    (user_id,),
                ).fetchone()
            else:
                plan_row = conn.execute("SELECT id FROM plans WHERE is_archived = 0 ORDER BY id ASC LIMIT 1").fetchone()
                if not plan_row:
                    plan_row = conn.execute("SELECT id FROM plans ORDER BY id ASC LIMIT 1").fetchone()

    if not plan_row:
        return None

    return get_plan_details(conn, plan_row["id"])


def update_plan(
    conn: sqlite3.Connection,
    plan_id: int,
    title: Optional[str] = None,
    description: Optional[str] = None,
    is_archived: Optional[bool] = None,
) -> Optional[Dict[str, Any]]:
    fields = []
    values = []
    if title is not None:
        fields.append("title = ?")
        values.append(title)
    if description is not None:
        fields.append("description = ?")
        values.append(description)
    if is_archived is not None:
        fields.append("is_archived = ?")
        values.append(1 if is_archived else 0)

    if fields:
        values.append(plan_id)
        cursor = conn.cursor()
        cursor.execute(f"UPDATE plans SET {', '.join(fields)} WHERE id = ?", tuple(values))
        conn.commit()
    return get_plan_details(conn, plan_id)


def get_all_plans(
    conn: sqlite3.Connection,
    user_id: Optional[int] = None,
    user_role: str = "admin",
    include_archived: bool = True,
) -> List[Dict[str, Any]]:
    if user_role == "admin" or user_id is None:
        if include_archived:
            rows = conn.execute("SELECT * FROM plans ORDER BY is_archived ASC, id ASC").fetchall()
        else:
            rows = conn.execute("SELECT * FROM plans WHERE is_archived = 0 ORDER BY id ASC").fetchall()
    else:
        # Check if user has explicit assignments
        has_assignments = conn.execute("SELECT 1 FROM user_plans WHERE user_id = ?", (user_id,)).fetchone()
        if has_assignments:
            if include_archived:
                rows = conn.execute(
                    """
                    SELECT p.* FROM plans p
                    JOIN user_plans up ON p.id = up.plan_id
                    WHERE up.user_id = ?
                    ORDER BY p.is_archived ASC, p.id ASC
                    """,
                    (user_id,),
                ).fetchall()
            else:
                rows = conn.execute(
                    """
                    SELECT p.* FROM plans p
                    JOIN user_plans up ON p.id = up.plan_id
                    WHERE up.user_id = ? AND p.is_archived = 0
                    ORDER BY p.id ASC
                    """,
                    (user_id,),
                ).fetchall()
        else:
            if include_archived:
                rows = conn.execute("SELECT * FROM plans ORDER BY is_archived ASC, id ASC").fetchall()
            else:
                rows = conn.execute("SELECT * FROM plans WHERE is_archived = 0 ORDER BY id ASC").fetchall()

    plans = []
    for r in rows:
        p = dict(r)
        p["is_archived"] = bool(p.get("is_archived", 0))
        ver_count = conn.execute("SELECT COUNT(*) as cnt FROM versions WHERE plan_id = ?", (p["id"],)).fetchone()["cnt"]
        p["versions_count"] = ver_count

        active_ver_row = conn.execute(
            "SELECT id, title FROM versions WHERE plan_id = ? AND is_active = 1 LIMIT 1",
            (p["id"],),
        ).fetchone()
        if not active_ver_row:
            active_ver_row = conn.execute(
                "SELECT id, title FROM versions WHERE plan_id = ? ORDER BY id DESC LIMIT 1",
                (p["id"],),
            ).fetchone()

        if active_ver_row:
            p["active_version_id"] = active_ver_row["id"]
            p["active_version_title"] = active_ver_row["title"]
            details = get_version_details(conn, active_ver_row["id"])
            if details and "totals" in details:
                p["total_expenses"] = details["totals"].get("total_expenses", 0.0)
                p["total_contributions"] = details["totals"].get("total_contributions", 0.0)
                p["total_balance"] = details["totals"].get("balance", 0.0)
            else:
                p["total_expenses"] = 0.0
                p["total_contributions"] = 0.0
                p["total_balance"] = 0.0
        else:
            p["active_version_id"] = None
            p["active_version_title"] = None
            p["total_expenses"] = 0.0
            p["total_contributions"] = 0.0
            p["total_balance"] = 0.0

        plans.append(p)
    return plans


def create_plan(conn: sqlite3.Connection, title: str, description: Optional[str] = None) -> Dict[str, Any]:
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO plans (title, description, is_archived) VALUES (?, ?, 0)",
        (title, description or ""),
    )
    plan_id = cursor.lastrowid
    # Create initial version
    cursor.execute(
        "INSERT INTO versions (plan_id, title, is_active, created_by) VALUES (?, ?, 1, 'Administrator')",
        (plan_id, "Aktueller Stand"),
    )
    conn.commit()
    details = get_plan_details(conn, plan_id)
    if details is None:
        raise ValueError("Fehler beim Erstellen des Plans")
    return details


def duplicate_plan(conn: sqlite3.Connection, plan_id: int, new_title: Optional[str] = None) -> Optional[Dict[str, Any]]:
    orig = conn.execute("SELECT * FROM plans WHERE id = ?", (plan_id,)).fetchone()
    if not orig:
        return None

    orig_dict = dict(orig)
    dup_title = new_title.strip() if new_title and new_title.strip() else f"{orig_dict['title']} (Kopie)"

    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO plans (title, description, is_archived) VALUES (?, ?, 0)",
        (dup_title, orig_dict.get("description") or ""),
    )
    new_plan_id = cursor.lastrowid
    if new_plan_id is None:
        raise ValueError("Fehler beim Duplizieren des Plans")

    # Fetch all versions of the original plan ordered by id asc
    ver_rows = conn.execute(
        "SELECT * FROM versions WHERE plan_id = ? ORDER BY id ASC",
        (plan_id,),
    ).fetchall()

    for v in ver_rows:
        v_dict = dict(v)
        cursor.execute(
            "INSERT INTO versions (plan_id, title, effective_date, is_active, created_by) VALUES (?, ?, ?, ?, ?)",
            (new_plan_id, v_dict["title"], v_dict.get("effective_date"), v_dict.get("is_active", 0), v_dict.get("created_by") or "Administrator"),
        )
        new_ver_id = cursor.lastrowid

        # Copy positions
        pos_rows = conn.execute("SELECT * FROM positions WHERE version_id = ?", (v_dict["id"],)).fetchall()
        for p in pos_rows:
            p_dict = dict(p)
            cursor.execute(
                "INSERT INTO positions (version_id, title, amount, comment, category, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
                (new_ver_id, p_dict["title"], p_dict["amount"], p_dict.get("comment"), p_dict.get("category"), p_dict.get("sort_order", 0)),
            )

        # Copy contributions
        contrib_rows = conn.execute("SELECT * FROM contributions WHERE version_id = ?", (v_dict["id"],)).fetchall()
        for c in contrib_rows:
            c_dict = dict(c)
            cursor.execute(
                "INSERT INTO contributions (version_id, person_name, amount, comment, sort_order) VALUES (?, ?, ?, ?, ?)",
                (new_ver_id, c_dict["person_name"], c_dict["amount"], c_dict.get("comment"), c_dict.get("sort_order", 0)),
            )

    conn.commit()
    return get_plan_details(conn, new_plan_id)


def delete_plan(conn: sqlite3.Connection, plan_id: int) -> bool:
    existing = conn.execute("SELECT * FROM plans WHERE id = ?", (plan_id,)).fetchone()
    if not existing:
        return False

    plan_count = conn.execute("SELECT COUNT(*) as cnt FROM plans").fetchone()["cnt"]
    if plan_count <= 1:
        raise ValueError("Der letzte verbleibende Plan kann nicht gelöscht werden. Mindestens ein Plan muss immer existieren.")

    cursor = conn.cursor()
    cursor.execute("DELETE FROM plans WHERE id = ?", (plan_id,))
    conn.commit()
    return True


def get_user_assigned_plans(conn: sqlite3.Connection, user_id: int) -> List[int]:
    rows = conn.execute("SELECT plan_id FROM user_plans WHERE user_id = ? ORDER BY plan_id ASC", (user_id,)).fetchall()
    return [r["plan_id"] for r in rows]


def set_user_assigned_plans(conn: sqlite3.Connection, user_id: int, plan_ids: List[int]):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM user_plans WHERE user_id = ?", (user_id,))
    for pid in plan_ids:
        cursor.execute("INSERT OR IGNORE INTO user_plans (user_id, plan_id) VALUES (?, ?)", (user_id, pid))
    conn.commit()


def check_user_plan_access(conn: sqlite3.Connection, user_id: int, user_role: str, plan_id: int) -> bool:
    if user_role == "admin":
        return True
    has_assignments = conn.execute("SELECT 1 FROM user_plans WHERE user_id = ?", (user_id,)).fetchone()
    if not has_assignments:
        return True  # Open access fallback when no specific restrictions configured
    allowed = conn.execute("SELECT 1 FROM user_plans WHERE user_id = ? AND plan_id = ?", (user_id, plan_id)).fetchone()
    return allowed is not None


def create_position(conn: sqlite3.Connection, version_id: int, title: str, amount: float, comment: Optional[str], category: Optional[str], sort_order: int = 0) -> Dict[str, Any]:
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO positions (version_id, title, amount, comment, category, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
        (version_id, title, amount, comment, category, sort_order),
    )
    conn.commit()
    pos_id = cursor.lastrowid
    row = conn.execute("SELECT * FROM positions WHERE id = ?", (pos_id,)).fetchone()
    res = dict(row)
    res["amount_formatted"] = format_currency_de(res["amount"])
    return res


def update_position(conn: sqlite3.Connection, position_id: int, title: Optional[str] = None, amount: Optional[float] = None, comment: Optional[str] = None, category: Optional[str] = None, sort_order: Optional[int] = None) -> Optional[Dict[str, Any]]:
    existing = conn.execute("SELECT * FROM positions WHERE id = ?", (position_id,)).fetchone()
    if not existing:
        return None

    current = dict(existing)
    new_title = title if title is not None else current["title"]
    new_amount = amount if amount is not None else current["amount"]
    new_comment = comment if comment is not None else current["comment"]
    new_category = category if category is not None else current["category"]
    new_sort = sort_order if sort_order is not None else current["sort_order"]

    conn.execute(
        "UPDATE positions SET title = ?, amount = ?, comment = ?, category = ?, sort_order = ? WHERE id = ?",
        (new_title, new_amount, new_comment, new_category, new_sort, position_id),
    )
    conn.commit()

    updated = conn.execute("SELECT * FROM positions WHERE id = ?", (position_id,)).fetchone()
    res = dict(updated)
    res["amount_formatted"] = format_currency_de(res["amount"])
    return res


def delete_position(conn: sqlite3.Connection, position_id: int) -> bool:
    cursor = conn.cursor()
    cursor.execute("DELETE FROM positions WHERE id = ?", (position_id,))
    conn.commit()
    return cursor.rowcount > 0


def create_contribution(conn: sqlite3.Connection, version_id: int, person_name: str, amount: float, comment: Optional[str], sort_order: int = 0) -> Dict[str, Any]:
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO contributions (version_id, person_name, amount, comment, sort_order) VALUES (?, ?, ?, ?, ?)",
        (version_id, person_name, amount, comment, sort_order),
    )
    conn.commit()
    c_id = cursor.lastrowid
    row = conn.execute("SELECT * FROM contributions WHERE id = ?", (c_id,)).fetchone()
    res = dict(row)
    res["amount_formatted"] = format_currency_de(res["amount"])
    return res


def update_contribution(conn: sqlite3.Connection, contribution_id: int, person_name: Optional[str] = None, amount: Optional[float] = None, comment: Optional[str] = None, sort_order: Optional[int] = None) -> Optional[Dict[str, Any]]:
    existing = conn.execute("SELECT * FROM contributions WHERE id = ?", (contribution_id,)).fetchone()
    if not existing:
        return None

    current = dict(existing)
    new_person = person_name if person_name is not None else current["person_name"]
    new_amount = amount if amount is not None else current["amount"]
    new_comment = comment if comment is not None else current["comment"]
    new_sort = sort_order if sort_order is not None else current["sort_order"]

    conn.execute(
        "UPDATE contributions SET person_name = ?, amount = ?, comment = ?, sort_order = ? WHERE id = ?",
        (new_person, new_amount, new_comment, new_sort, contribution_id),
    )
    conn.commit()

    updated = conn.execute("SELECT * FROM contributions WHERE id = ?", (contribution_id,)).fetchone()
    res = dict(updated)
    res["amount_formatted"] = format_currency_de(res["amount"])
    return res


def delete_contribution(conn: sqlite3.Connection, contribution_id: int) -> bool:
    cursor = conn.cursor()
    cursor.execute("DELETE FROM contributions WHERE id = ?", (contribution_id,))
    conn.commit()
    return cursor.rowcount > 0


def save_new_version(
    conn: sqlite3.Connection,
    plan_id: int,
    title: str,
    effective_date: Optional[str] = None,
    positions: Optional[List[Dict[str, Any]]] = None,
    contributions: Optional[List[Dict[str, Any]]] = None,
    created_by: Optional[str] = "Administrator",
) -> Dict[str, Any]:
    cursor = conn.cursor()
    # Deactivate existing active versions for this plan
    cursor.execute("UPDATE versions SET is_active = 0 WHERE plan_id = ?", (plan_id,))

    # Insert new active version with created_by audit metadata
    cursor.execute(
        "INSERT INTO versions (plan_id, title, effective_date, is_active, created_by) VALUES (?, ?, ?, 1, ?)",
        (plan_id, title, effective_date, created_by or "Administrator"),
    )
    new_ver_id = cursor.lastrowid
    if new_ver_id is None:
        raise ValueError("Fehler beim Erstellen der Version")

    # Insert positions
    if positions:
        for idx, pos in enumerate(positions):
            p_title = pos.get("title") or ""
            p_amount = float(pos.get("amount") or 0.0)
            p_comment = pos.get("comment")
            p_cat = pos.get("category")
            p_sort = pos.get("sort_order", idx)
            cursor.execute(
                "INSERT INTO positions (version_id, title, amount, comment, category, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
                (new_ver_id, p_title, p_amount, p_comment, p_cat, p_sort),
            )

    # Insert contributions
    if contributions:
        for idx, c in enumerate(contributions):
            c_person = c.get("person_name") or ""
            c_amount = float(c.get("amount") or 0.0)
            c_comment = c.get("comment")
            c_sort = c.get("sort_order", idx)
            cursor.execute(
                "INSERT INTO contributions (version_id, person_name, amount, comment, sort_order) VALUES (?, ?, ?, ?, ?)",
                (new_ver_id, c_person, c_amount, c_comment, c_sort),
            )

    conn.commit()
    details = get_version_details(conn, new_ver_id)
    if details is None:
        raise ValueError("Fehler beim Abrufen der erstellten Version")
    return details


def update_version(
    conn: sqlite3.Connection,
    version_id: int,
    title: Optional[str] = None,
    effective_date: Optional[str] = None,
    updated_by: Optional[str] = "Administrator",
) -> Optional[Dict[str, Any]]:
    existing = conn.execute("SELECT * FROM versions WHERE id = ?", (version_id,)).fetchone()
    if not existing:
        return None

    current = dict(existing)
    new_title = title if title is not None else current["title"]
    new_date = effective_date if effective_date is not None else current["effective_date"]
    now_iso = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn.execute(
        "UPDATE versions SET title = ?, effective_date = ?, updated_at = ?, updated_by = ? WHERE id = ?",
        (new_title, new_date, now_iso, updated_by or "Administrator", version_id),
    )
    conn.commit()
    return get_version_details(conn, version_id)


def delete_version(conn: sqlite3.Connection, version_id: int) -> bool:
    ver_row = conn.execute("SELECT * FROM versions WHERE id = ?", (version_id,)).fetchone()
    if not ver_row:
        return False

    ver = dict(ver_row)
    plan_id = ver["plan_id"]
    is_active = ver["is_active"]

    count_row = conn.execute("SELECT COUNT(*) as cnt FROM versions WHERE plan_id = ?", (plan_id,)).fetchone()
    if count_row and count_row["cnt"] <= 1:
        raise ValueError("Der letzte verbleibende Stand eines Plans kann nicht gelöscht werden.")

    cursor = conn.cursor()
    cursor.execute("DELETE FROM positions WHERE version_id = ?", (version_id,))
    cursor.execute("DELETE FROM contributions WHERE version_id = ?", (version_id,))
    cursor.execute("DELETE FROM versions WHERE id = ?", (version_id,))

    # If the deleted version was active, make the latest remaining version active
    if is_active == 1:
        latest = conn.execute("SELECT id FROM versions WHERE plan_id = ? ORDER BY id DESC LIMIT 1", (plan_id,)).fetchone()
        if latest:
            cursor.execute("UPDATE versions SET is_active = 1 WHERE id = ?", (latest["id"],))

    conn.commit()
    return True


def activate_version(conn: sqlite3.Connection, plan_id: int, version_id: int) -> Optional[Dict[str, Any]]:
    ver = conn.execute("SELECT id FROM versions WHERE id = ? AND plan_id = ?", (version_id, plan_id)).fetchone()
    if not ver:
        return None
    cursor = conn.cursor()
    cursor.execute("UPDATE versions SET is_active = 0 WHERE plan_id = ?", (plan_id,))
    cursor.execute("UPDATE versions SET is_active = 1 WHERE id = ?", (version_id,))
    conn.commit()
    return get_version_details(conn, version_id)


def get_plan_history(conn: sqlite3.Connection, plan_id: int) -> List[Dict[str, Any]]:
    ver_rows = conn.execute(
        "SELECT id, plan_id, title, effective_date, is_active, created_at, created_by, updated_at, updated_by FROM versions WHERE plan_id = ? ORDER BY id DESC",
        (plan_id,),
    ).fetchall()

    history = []
    for r in ver_rows:
        v_dict = dict(r)
        v_id = v_dict["id"]
        v_details = get_version_details(conn, v_id)
        if v_details:
            v_dict["positions_count"] = len(v_details["positions"])
            v_dict["contributions_count"] = len(v_details["contributions"])
            v_dict["totals"] = v_details["totals"]
        else:
            v_dict["positions_count"] = 0
            v_dict["contributions_count"] = 0
            v_dict["totals"] = {}
        history.append(v_dict)

    return history


def create_version_snapshot(conn: sqlite3.Connection, plan_id: int, title: str, effective_date: Optional[str], copy_from_version_id: Optional[int] = None) -> Dict[str, Any]:
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO versions (plan_id, title, effective_date, is_active) VALUES (?, ?, ?, 1)",
        (plan_id, title, effective_date),
    )
    new_ver_id = cursor.lastrowid
    if new_ver_id is None:
        raise ValueError("Fehler beim Erstellen der Snapshot-Version")

    # If copying from an existing version, duplicate positions & contributions
    if copy_from_version_id:
        positions = conn.execute("SELECT title, amount, comment, category, sort_order FROM positions WHERE version_id = ?", (copy_from_version_id,)).fetchall()
        for p in positions:
            cursor.execute(
                "INSERT INTO positions (version_id, title, amount, comment, category, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
                (new_ver_id, p["title"], p["amount"], p["comment"], p["category"], p["sort_order"]),
            )

        contributions = conn.execute("SELECT person_name, amount, comment, sort_order FROM contributions WHERE version_id = ?", (copy_from_version_id,)).fetchall()
        for c in contributions:
            cursor.execute(
                "INSERT INTO contributions (version_id, person_name, amount, comment, sort_order) VALUES (?, ?, ?, ?, ?)",
                (new_ver_id, c["person_name"], c["amount"], c["comment"], c["sort_order"]),
            )

    conn.commit()
    details = get_version_details(conn, new_ver_id)
    if details is None:
        raise ValueError("Fehler beim Abrufen der Snapshot-Version")
    return details


def get_history_comparison(conn: sqlite3.Connection, plan_id: int) -> Dict[str, Any]:
    ver_rows = conn.execute(
        "SELECT id, title, effective_date, created_at, created_by, updated_at, updated_by FROM versions WHERE plan_id = ? ORDER BY id ASC",
        (plan_id,),
    ).fetchall()

    versions = [dict(v) for v in ver_rows]

    # Collect all positions across versions
    positions_map: Dict[str, Dict[str, Any]] = {}  # title -> { category, comment, values: {v_id: amount} }
    # Collect all contributions across versions
    contributions_map: Dict[str, Dict[str, Any]] = {}  # person_name -> { comment, values: {v_id: amount} }

    totals_by_version: Dict[str, Dict[str, Any]] = {}

    for v in versions:
        v_id = v["id"]
        v_details = get_version_details(conn, v_id)
        if not v_details:
            continue
        totals_by_version[str(v_id)] = v_details["totals"]

        for pos in v_details["positions"]:
            title = pos["title"]
            if title not in positions_map:
                positions_map[title] = {
                    "title": title,
                    "category": pos.get("category"),
                    "comment": pos.get("comment"),
                    "values": {},
                    "formatted_values": {},
                }
            positions_map[title]["values"][str(v_id)] = pos["amount"]
            positions_map[title]["formatted_values"][str(v_id)] = pos["amount_formatted"]

        for c in v_details["contributions"]:
            person = c["person_name"]
            if person not in contributions_map:
                contributions_map[person] = {
                    "title": f"Zahlung {person}",
                    "comment": c.get("comment"),
                    "values": {},
                    "formatted_values": {},
                }
            contributions_map[person]["values"][str(v_id)] = c["amount"]
            contributions_map[person]["formatted_values"][str(v_id)] = c["amount_formatted"]


    return {
        "versions": versions,
        "rows": list(positions_map.values()),
        "contributions_rows": list(contributions_map.values()),
        "totals": totals_by_version,
    }


def export_full_data(conn: sqlite3.Connection) -> Dict[str, Any]:
    plan_rows = conn.execute("SELECT * FROM plans ORDER BY id ASC").fetchall()
    plans = []

    for p in plan_rows:
        p_dict = dict(p)
        plan_id = p_dict["id"]

        ver_rows = conn.execute("SELECT * FROM versions WHERE plan_id = ? ORDER BY id ASC", (plan_id,)).fetchall()
        versions = []

        for v in ver_rows:
            v_dict = dict(v)
            v_id = v_dict["id"]

            pos_rows = conn.execute(
                "SELECT title, amount, comment, category, sort_order FROM positions WHERE version_id = ? ORDER BY sort_order ASC, id ASC",
                (v_id,),
            ).fetchall()
            positions = [dict(pos) for pos in pos_rows]

            contrib_rows = conn.execute(
                "SELECT person_name, amount, comment, sort_order FROM contributions WHERE version_id = ? ORDER BY sort_order ASC, id ASC",
                (v_id,),
            ).fetchall()
            contributions = [dict(c) for c in contrib_rows]

            versions.append(
                {
                    "title": v_dict["title"],
                    "effective_date": v_dict["effective_date"],
                    "is_active": v_dict["is_active"],
                    "created_at": v_dict.get("created_at"),
                    "created_by": v_dict.get("created_by"),
                    "updated_at": v_dict.get("updated_at"),
                    "updated_by": v_dict.get("updated_by"),
                    "positions": positions,
                    "contributions": contributions,
                }
            )

        plans.append(
            {
                "title": p_dict["title"],
                "description": p_dict["description"],
                "versions": versions,
            }
        )

    return {
        "version": 1,
        "exported_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "plans": plans,
    }


def export_full_data_xlsx(conn: sqlite3.Connection) -> bytes:
    """Export complete plan, active version details and historical snapshots to an Excel (.xlsx) workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_section_header = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_kpi_num = Font(name="Calibri", size=14, bold=True)

    fill_navy = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_indigo = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    fill_emerald = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    fill_slate_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_kpi_box = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    double_border_bottom = Side(border_style="double", color="334155")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    total_border = Border(top=thin_border_side, bottom=double_border_bottom)

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    currency_fmt = '#,##0.00\\ \\€'

    full_data = export_full_data(conn)
    plans = full_data.get("plans", [])

    for p_idx, plan in enumerate(plans):
        plan_title = plan.get("title", f"Plan {p_idx+1}")
        versions = plan.get("versions", [])

        # Find active version or latest
        active_ver = next((v for v in versions if v.get("is_active")), versions[-1] if versions else None)

        if active_ver:
            # Sheet 1: Active Overview
            sheet_title = f"Aktuell - {plan_title}"[:31].replace(":", "-").replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "")
            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            # Plan Title & Subtitle
            cell_a1 = ws.cell(row=1, column=1, value=f"📊 Ausgabenplaner: {plan_title}")
            cell_a1.font = font_title
            cell_a2 = ws.cell(row=2, column=1, value=f"Stand: {active_ver.get('title')} | Gültig ab: {active_ver.get('effective_date') or '-'} | Exportiert am: {get_local_now().strftime('%d.%m.%Y %H:%M')}")
            cell_a2.font = font_subtitle

            # Section: Positions (Ausgaben)
            cur_row = 4
            ws.cell(row=cur_row, column=1, value="💸 Monatliche Ausgaben (Positionen)").font = font_section_header
            ws.cell(row=cur_row, column=1).fill = fill_indigo
            for col in range(2, 6):
                ws.cell(row=cur_row, column=col).fill = fill_indigo
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)

            cur_row += 1
            pos_headers = ["Nr.", "Position / Ausgabezweck", "Kategorie", "Monatlicher Betrag (€)", "Kommentar / Intervall"]
            for c_idx, h in enumerate(pos_headers, 1):
                c = ws.cell(row=cur_row, column=c_idx, value=h)
                c.font = font_tbl_header
                c.fill = fill_slate_header
                c.alignment = align_center if c_idx == 1 else (align_right if c_idx == 4 else align_left)
                c.border = cell_border

            pos_start_row = cur_row + 1
            positions = active_ver.get("positions", [])
            for idx, pos in enumerate(positions, 1):
                cur_row += 1
                r_fill = fill_zebra if idx % 2 == 0 else None

                c1 = ws.cell(row=cur_row, column=1, value=idx)
                c1.alignment = align_center
                c1.font = font_regular
                c1.border = cell_border
                if r_fill: c1.fill = r_fill

                c2 = ws.cell(row=cur_row, column=2, value=pos.get("title", ""))
                c2.alignment = align_left
                c2.font = font_regular
                c2.border = cell_border
                if r_fill: c2.fill = r_fill

                c3 = ws.cell(row=cur_row, column=3, value=pos.get("category", "") or "-")
                c3.alignment = align_left
                c3.font = font_regular
                c3.border = cell_border
                if r_fill: c3.fill = r_fill

                c4 = ws.cell(row=cur_row, column=4, value=float(pos.get("amount", 0.0)))
                c4.alignment = align_right
                c4.font = font_regular
                c4.number_format = currency_fmt
                c4.border = cell_border
                if r_fill: c4.fill = r_fill

                c5 = ws.cell(row=cur_row, column=5, value=pos.get("comment", "") or "")
                c5.alignment = align_left
                c5.font = font_regular
                c5.border = cell_border
                if r_fill: c5.fill = r_fill

            # Total row for Positions
            cur_row += 1
            ws.cell(row=cur_row, column=1, value="").border = total_border
            ws.cell(row=cur_row, column=1).fill = fill_total
            c_pos_label = ws.cell(row=cur_row, column=2, value="Gesamtausgaben:")
            c_pos_label.font = font_bold
            c_pos_label.border = total_border
            c_pos_label.fill = fill_total

            ws.cell(row=cur_row, column=3, value="").border = total_border
            ws.cell(row=cur_row, column=3).fill = fill_total

            c_pos_sum = ws.cell(row=cur_row, column=4)
            if positions:
                c_pos_sum.value = f"=SUM(D{pos_start_row}:D{cur_row-1})"
            else:
                c_pos_sum.value = 0.0
            c_pos_sum.font = font_bold
            c_pos_sum.alignment = align_right
            c_pos_sum.number_format = currency_fmt
            c_pos_sum.border = total_border
            c_pos_sum.fill = fill_total
            pos_sum_cell_ref = f"D{cur_row}"

            ws.cell(row=cur_row, column=5, value="").border = total_border
            ws.cell(row=cur_row, column=5).fill = fill_total

            # Section: Contributions (Beiträge)
            cur_row += 3
            ws.cell(row=cur_row, column=1, value="💰 Monatliche Beiträge & Einnahmen").font = font_section_header
            ws.cell(row=cur_row, column=1).fill = fill_emerald
            for col in range(2, 5):
                ws.cell(row=cur_row, column=col).fill = fill_emerald
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=4)

            cur_row += 1
            contrib_headers = ["Nr.", "Person / Quelle", "Monatlicher Beitrag (€)", "Notiz / Kommentar"]
            for c_idx, h in enumerate(contrib_headers, 1):
                c = ws.cell(row=cur_row, column=c_idx, value=h)
                c.font = font_tbl_header
                c.fill = fill_slate_header
                c.alignment = align_center if c_idx == 1 else (align_right if c_idx == 3 else align_left)
                c.border = cell_border

            contrib_start_row = cur_row + 1
            contributions = active_ver.get("contributions", [])
            for idx, con in enumerate(contributions, 1):
                cur_row += 1
                r_fill = fill_zebra if idx % 2 == 0 else None

                c1 = ws.cell(row=cur_row, column=1, value=idx)
                c1.alignment = align_center
                c1.font = font_regular
                c1.border = cell_border
                if r_fill: c1.fill = r_fill

                c2 = ws.cell(row=cur_row, column=2, value=con.get("person_name", ""))
                c2.alignment = align_left
                c2.font = font_regular
                c2.border = cell_border
                if r_fill: c2.fill = r_fill

                c3 = ws.cell(row=cur_row, column=3, value=float(con.get("amount", 0.0)))
                c3.alignment = align_right
                c3.font = font_regular
                c3.number_format = currency_fmt
                c3.border = cell_border
                if r_fill: c3.fill = r_fill

                c4 = ws.cell(row=cur_row, column=4, value=con.get("comment", "") or "")
                c4.alignment = align_left
                c4.font = font_regular
                c4.border = cell_border
                if r_fill: c4.fill = r_fill

            # Total row for Contributions
            cur_row += 1
            ws.cell(row=cur_row, column=1, value="").border = total_border
            ws.cell(row=cur_row, column=1).fill = fill_total
            c_con_label = ws.cell(row=cur_row, column=2, value="Gesamtbeiträge:")
            c_con_label.font = font_bold
            c_con_label.border = total_border
            c_con_label.fill = fill_total

            c_con_sum = ws.cell(row=cur_row, column=3)
            if contributions:
                c_con_sum.value = f"=SUM(C{contrib_start_row}:C{cur_row-1})"
            else:
                c_con_sum.value = 0.0
            c_con_sum.font = font_bold
            c_con_sum.alignment = align_right
            c_con_sum.number_format = currency_fmt
            c_con_sum.border = total_border
            c_con_sum.fill = fill_total
            con_sum_cell_ref = f"C{cur_row}"

            ws.cell(row=cur_row, column=4, value="").border = total_border
            ws.cell(row=cur_row, column=4).fill = fill_total

            # Summary Saldo Block
            cur_row += 3
            ws.cell(row=cur_row, column=1, value="📈 Zusammenfassung & Saldo").font = font_section_header
            ws.cell(row=cur_row, column=1).fill = fill_navy
            for col in range(2, 4):
                ws.cell(row=cur_row, column=col).fill = fill_navy
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)

            cur_row += 1
            ws.cell(row=cur_row, column=1, value="Gesamtausgaben").font = font_regular
            ws.cell(row=cur_row, column=1).border = cell_border
            ws.cell(row=cur_row, column=2, value="").border = cell_border
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
            c_tot_exp = ws.cell(row=cur_row, column=3, value=f"={pos_sum_cell_ref}")
            c_tot_exp.font = font_bold
            c_tot_exp.alignment = align_right
            c_tot_exp.number_format = currency_fmt
            c_tot_exp.border = cell_border

            cur_row += 1
            ws.cell(row=cur_row, column=1, value="Gesamtbeiträge").font = font_regular
            ws.cell(row=cur_row, column=1).border = cell_border
            ws.cell(row=cur_row, column=2, value="").border = cell_border
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
            c_tot_inc = ws.cell(row=cur_row, column=3, value=f"={con_sum_cell_ref}")
            c_tot_inc.font = font_bold
            c_tot_inc.alignment = align_right
            c_tot_inc.number_format = currency_fmt
            c_tot_inc.border = cell_border

            cur_row += 1
            c_sal_lbl = ws.cell(row=cur_row, column=1, value="Saldo / Differenz")
            c_sal_lbl.font = font_bold
            c_sal_lbl.fill = fill_kpi_box
            c_sal_lbl.border = total_border
            ws.cell(row=cur_row, column=2, value="").fill = fill_kpi_box
            ws.cell(row=cur_row, column=2).border = total_border
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)

            c_sal_val = ws.cell(row=cur_row, column=3, value=f"=C{cur_row-1}+C{cur_row-2}")
            c_sal_val.font = font_kpi_num
            c_sal_val.alignment = align_right
            c_sal_val.number_format = currency_fmt
            c_sal_val.fill = fill_kpi_box
            c_sal_val.border = total_border

            # Adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if not val_str.startswith("="):
                            max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 36)
            ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width, 24)
            ws.column_dimensions["E"].width = max(ws.column_dimensions["E"].width, 32)

    # Sheet 2: Versions History (Alle Stände)
    ws_hist = wb.create_sheet(title="Historie & Stände")
    ws_hist.views.sheetView[0].showGridLines = True

    cell_h1 = ws_hist.cell(row=1, column=1, value="📜 Historienübersicht aller Versionen & Stände")
    cell_h1.font = font_title
    cell_h2 = ws_hist.cell(row=2, column=1, value=f"Exportiert am: {get_local_now().strftime('%d.%m.%Y %H:%M')}")
    cell_h2.font = font_subtitle

    hist_headers = ["Plan", "Stand-Name", "Gültig ab", "Status", "Erstellt von", "Erstellt am", "Geändert von", "Geändert am", "Ausgaben (€)", "Beiträge (€)", "Saldo (€)"]
    h_row = 4
    for c_idx, h in enumerate(hist_headers, 1):
        c = ws_hist.cell(row=h_row, column=c_idx, value=h)
        c.font = font_tbl_header
        c.fill = fill_navy
        c.alignment = align_center if c_idx in [3, 4, 6, 8] else (align_right if c_idx in [9, 10, 11] else align_left)
        c.border = cell_border

    row_idx = h_row
    item_count = 0
    for plan in plans:
        plan_title = plan.get("title", "")
        for v in plan.get("versions", []):
            row_idx += 1
            item_count += 1
            r_fill = fill_zebra if item_count % 2 == 0 else None

            v_positions = v.get("positions", [])
            v_contributions = v.get("contributions", [])
            exp_sum = sum(float(p.get("amount", 0.0)) for p in v_positions)
            inc_sum = sum(float(c.get("amount", 0.0)) for c in v_contributions)
            saldo = exp_sum + inc_sum

            vals = [
                (plan_title, align_left, font_regular, None),
                (v.get("title", ""), align_left, font_bold, None),
                (v.get("effective_date") or "-", align_center, font_regular, None),
                ("Aktiv" if v.get("is_active") else "Historisch", align_center, font_regular, None),
                (v.get("created_by") or "-", align_left, font_regular, None),
                (v.get("created_at") or "-", align_center, font_regular, None),
                (v.get("updated_by") or "-", align_left, font_regular, None),
                (v.get("updated_at") or "-", align_center, font_regular, None),
                (exp_sum, align_right, font_regular, currency_fmt),
                (inc_sum, align_right, font_regular, currency_fmt),
                (saldo, align_right, font_bold, currency_fmt),
            ]

            for c_idx, (val, align, fnt, num_fmt) in enumerate(vals, 1):
                c = ws_hist.cell(row=row_idx, column=c_idx, value=val)
                c.alignment = align
                c.font = fnt
                c.border = cell_border
                if num_fmt:
                    c.number_format = num_fmt
                if r_fill:
                    c.fill = r_fill

    # Adjust hist columns
    for col in ws_hist.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str))
        ws_hist.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def import_full_data(conn: sqlite3.Connection, data: Dict[str, Any], overwrite: bool = True) -> Dict[str, Any]:
    cursor = conn.cursor()

    if overwrite:
        cursor.execute("DELETE FROM contributions;")
        cursor.execute("DELETE FROM positions;")
        cursor.execute("DELETE FROM versions;")
        cursor.execute("DELETE FROM plans;")
        conn.commit()

    plans_imported = 0
    versions_imported = 0
    positions_imported = 0
    contributions_imported = 0

    plans = data.get("plans", [])
    for p in plans:
        cursor.execute("INSERT INTO plans (title, description) VALUES (?, ?)", (p["title"], p.get("description")))
        plan_id = cursor.lastrowid
        plans_imported += 1

        versions = p.get("versions", [])
        for v in versions:
            cursor.execute(
                "INSERT INTO versions (plan_id, title, effective_date, is_active, created_at, created_by, updated_at, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    plan_id,
                    v["title"],
                    v.get("effective_date"),
                    v.get("is_active", 1),
                    v.get("created_at") or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    v.get("created_by") or "Administrator",
                    v.get("updated_at"),
                    v.get("updated_by"),
                ),
            )
            version_id = cursor.lastrowid
            versions_imported += 1

            positions = v.get("positions", [])
            for idx, pos in enumerate(positions):
                cursor.execute(
                    "INSERT INTO positions (version_id, title, amount, comment, category, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
                    (version_id, pos["title"], pos["amount"], pos.get("comment"), pos.get("category"), pos.get("sort_order", idx)),
                )
                positions_imported += 1

            contributions = v.get("contributions", [])
            for idx, c in enumerate(contributions):
                cursor.execute(
                    "INSERT INTO contributions (version_id, person_name, amount, comment, sort_order) VALUES (?, ?, ?, ?, ?)",
                    (version_id, c["person_name"], c["amount"], c.get("comment"), c.get("sort_order", idx)),
                )
                contributions_imported += 1

    conn.commit()
    return {
        "success": True,
        "plans_imported": plans_imported,
        "versions_imported": versions_imported,
        "positions_imported": positions_imported,
        "contributions_imported": contributions_imported,
    }

