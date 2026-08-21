import pytest
from fastapi.testclient import TestClient
from app.main import app
from app.database import init_db, reset_db

client = TestClient(app)


@pytest.fixture(autouse=True)
def setup_test_db():
    reset_db()
    init_db()
    yield


def test_export_and_import_flow():
    # Login admin
    login_resp = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = login_resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Add sample position & contribution to active plan
    plan_resp = client.get("/api/plans/active", headers=headers)
    assert plan_resp.status_code == 200
    version_id = plan_resp.json()["active_version"]["id"]

    pos_resp = client.post(
        f"/api/versions/{version_id}/positions",
        json={"title": "Miete Export Test", "amount": -1000.0, "comment": "Test Comment", "category": "Wohnen"},
        headers=headers,
    )
    assert pos_resp.status_code == 201

    contrib_resp = client.post(
        f"/api/versions/{version_id}/contributions",
        json={"person_name": "Alex", "amount": 500.0, "comment": "Alex Beitrag"},
        headers=headers,
    )
    assert contrib_resp.status_code == 201

    # 1. Export data
    export_resp = client.get("/api/data/export", headers=headers)
    assert export_resp.status_code == 200
    export_data = export_resp.json()

    assert export_data["version"] == 1
    assert len(export_data["plans"]) == 1
    plan_data = export_data["plans"][0]
    assert plan_data["title"] == "Muster-Wirtschaftsplan"
    assert len(plan_data["versions"]) == 1

    ver_data = plan_data["versions"][0]
    assert len(ver_data["positions"]) == 1
    assert ver_data["positions"][0]["title"] == "Miete Export Test"
    assert ver_data["positions"][0]["amount"] == -1000.0

    assert len(ver_data["contributions"]) == 1
    assert ver_data["contributions"][0]["person_name"] == "Alex"
    assert ver_data["contributions"][0]["amount"] == 500.0

    # 2. Clear database
    reset_db()
    init_db(seed=False)

    # Confirm DB has 0 positions and 0 contributions
    ver_check = client.get(f"/api/versions/{version_id}", headers=headers)
    assert ver_check.status_code == 200
    assert len(ver_check.json()["positions"]) == 0
    assert len(ver_check.json()["contributions"]) == 0

    # 3. Import data back
    import_resp = client.post("/api/data/import", json=export_data, headers=headers)
    assert import_resp.status_code == 200
    import_res = import_resp.json()
    assert import_res["success"] is True
    assert import_res["positions_imported"] == 1
    assert import_res["contributions_imported"] == 1

    # 4. Verify restored active plan data
    restored_plan = client.get("/api/plans/active", headers=headers).json()
    restored_ver = restored_plan["active_version"]
    assert len(restored_ver["positions"]) == 1
    assert restored_ver["positions"][0]["title"] == "Miete Export Test"
    assert restored_ver["positions"][0]["amount"] == -1000.0

    assert len(restored_ver["contributions"]) == 1
    assert restored_ver["contributions"][0]["person_name"] == "Alex"
    assert restored_ver["contributions"][0]["amount"] == 500.0
    assert restored_ver["totals"]["net_balance"] == -500.0


def test_export_includes_entire_history_and_restores_all_versions():
    # Login admin
    login_resp = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    plan_resp = client.get("/api/plans/active", headers=headers).json()
    plan_id = plan_resp["id"]

    # Save Version 2 (Stand 01.10.2026)
    client.post(
        f"/api/plans/{plan_id}/save-version",
        json={
            "title": "Stand ab 01.10.2026",
            "effective_date": "2026-10-01",
            "positions": [{"title": "Strom", "amount": -80.0}],
            "contributions": [{"person_name": "Person A", "amount": 80.0}],
        },
        headers=headers,
    )

    # Save Version 3 (Stand 01.11.2026)
    client.post(
        f"/api/plans/{plan_id}/save-version",
        json={
            "title": "Stand ab 01.11.2026",
            "effective_date": "2026-11-01",
            "positions": [{"title": "Gas", "amount": -120.0}],
            "contributions": [{"person_name": "Person B", "amount": 120.0}],
        },
        headers=headers,
    )

    # Export
    export_resp = client.get("/api/data/export", headers=headers)
    assert export_resp.status_code == 200
    export_data = export_resp.json()

    # Verify all 3 historical versions are included in export
    versions = export_data["plans"][0]["versions"]
    assert len(versions) == 3
    version_titles = [v["title"] for v in versions]
    assert "Aktueller Stand" in version_titles
    assert "Stand ab 01.10.2026" in version_titles
    assert "Stand ab 01.11.2026" in version_titles

    # Reset DB & restore
    reset_db()
    init_db(seed=False)

    import_resp = client.post("/api/data/import", json=export_data, headers=headers)
    assert import_resp.status_code == 200
    assert import_resp.json()["versions_imported"] == 3

    # Check restored history
    restored_plan = client.get("/api/plans/active", headers=headers).json()
    restored_plan_id = restored_plan["id"]
    history_resp = client.get(f"/api/plans/{restored_plan_id}/history", headers=headers)
    assert history_resp.status_code == 200
    restored_history = history_resp.json()
    assert len(restored_history) == 3


def test_export_xlsx_endpoint_and_structure():
    import io
    import openpyxl

    # Login
    login_resp = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = login_resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Populate active version with data
    plan_resp = client.get("/api/plans/active", headers=headers)
    plan_data = plan_resp.json()
    version_id = plan_data["active_version"]["id"]

    client.post(
        f"/api/versions/{version_id}/positions",
        json={"title": "Miete XLSX Test", "amount": -1050.0, "comment": "Kaltmiete", "category": "Wohnen"},
        headers=headers,
    )
    client.post(
        f"/api/versions/{version_id}/contributions",
        json={"person_name": "Alex", "amount": 600.0, "comment": "Anteil Alex"},
        headers=headers,
    )

    # 1. Test GET /api/data/export-xlsx
    export_resp = client.get("/api/data/export-xlsx", headers=headers)
    assert export_resp.status_code == 200
    assert "spreadsheetml" in export_resp.headers["content-type"]
    assert "ausgabenplaner_export_" in export_resp.headers["content-disposition"]
    assert len(export_resp.content) > 0

    # 2. Parse workbook with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
    sheet_names = wb.sheetnames
    assert len(sheet_names) >= 2
    assert "Historie & Stände" in sheet_names

    # Check active plan sheet
    active_sheet = wb[sheet_names[0]]
    assert "Ausgabenplaner" in str(active_sheet["A1"].value)

    # Search for position and contribution values in active sheet
    cell_values = [str(cell.value) for row in active_sheet.iter_rows() for cell in row if cell.value is not None]
    assert any("Miete XLSX Test" in v for v in cell_values)
    assert any("Alex" in v for v in cell_values)
    assert any("Saldo" in v for v in cell_values)

    # Check history sheet
    hist_sheet = wb["Historie & Stände"]
    hist_values = [str(cell.value) for row in hist_sheet.iter_rows() for cell in row if cell.value is not None]
    assert "Plan" in hist_values
    assert "Stand-Name" in hist_values
    assert "Saldo (€)" in hist_values


def test_export_import_custom_categories():
    # Login admin
    login_resp = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = login_resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Create a custom category
    cat_resp = client.post(
        "/api/categories",
        json={"name": "Haustier & Tierarzt", "color": "#f43f5e", "icon": "🐱", "sort_order": 99},
        headers=headers,
    )
    assert cat_resp.status_code == 201

    # Export
    export_resp = client.get("/api/data/export", headers=headers)
    assert export_resp.status_code == 200
    export_data = export_resp.json()

    assert "categories" in export_data
    assert len(export_data["categories"]) > 0
    custom_cat = next((c for c in export_data["categories"] if c["name"] == "Haustier & Tierarzt"), None)
    assert custom_cat is not None
    assert custom_cat["color"] == "#f43f5e"
    assert custom_cat["icon"] == "🐱"
    assert custom_cat["sort_order"] == 99

    # Reset DB (without seeding default categories)
    reset_db()
    init_db(seed=False)

    # Verify categories are empty / default only
    cats_before = client.get("/api/categories", headers=headers).json()
    assert not any(c["name"] == "Haustier & Tierarzt" for c in cats_before)

    # Import
    import_resp = client.post("/api/data/import", json=export_data, headers=headers)
    assert import_resp.status_code == 200
    assert import_resp.json()["success"] is True

    # Verify categories are restored
    cats_after = client.get("/api/categories", headers=headers).json()
    restored_cat = next((c for c in cats_after if c["name"] == "Haustier & Tierarzt"), None)
    assert restored_cat is not None
    assert restored_cat["color"] == "#f43f5e"
    assert restored_cat["icon"] == "🐱"
    assert restored_cat["sort_order"] == 99


def test_export_import_preserves_archived_plan_state():
    # Login admin
    login_resp = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    token = login_resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Create plan 1 (active)
    p1 = client.post("/api/plans", json={"title": "Aktiver Plan 2026", "description": "Laufend"}, headers=headers).json()
    # Create plan 2 (archived)
    p2 = client.post("/api/plans", json={"title": "Altplan 2025", "description": "Abgeschlossen"}, headers=headers).json()
    patch_resp = client.patch(f"/api/plans/{p2['id']}", json={"is_archived": True}, headers=headers)
    assert patch_resp.status_code == 200

    # Export
    export_resp = client.get("/api/data/export", headers=headers)
    assert export_resp.status_code == 200
    export_data = export_resp.json()

    exported_plans = export_data["plans"]
    p1_exp = next(p for p in exported_plans if p["title"] == "Aktiver Plan 2026")
    p2_exp = next(p for p in exported_plans if p["title"] == "Altplan 2025")
    assert p1_exp["is_archived"] is False
    assert p2_exp["is_archived"] is True

    # Reset DB & re-import
    reset_db()
    init_db(seed=False)

    import_resp = client.post("/api/data/import", json=export_data, headers=headers)
    assert import_resp.status_code == 200

    summaries = client.get("/api/plans?include_archived=true", headers=headers).json()
    p1_restored = next(p for p in summaries if p["title"] == "Aktiver Plan 2026")
    p2_restored = next(p for p in summaries if p["title"] == "Altplan 2025")
    assert p1_restored["is_archived"] is False
    assert p2_restored["is_archived"] is True



